from django.conf import settings
from rest_framework.exceptions import APIException

from openpyxl import load_workbook, worksheet


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def _clean_merge_range(self, cr):
        """
        Remove all but the top left-cell from a range of merged cells
        """

        min_col, min_row, max_col, max_row = cr.bounds
        rows = range(min_row, max_row + 1)
        cols = range(min_col, max_col + 1)
        # cells = product(rows, cols)

        # for c in islice(cells, 1, None):
        #    if c in self._cells:
        #        del self._cells[c]

    # Apply monkey patch
    worksheet.Worksheet._clean_merge_range = _clean_merge_range


class ExcelGenerator:
    def __init__(self, template_name, sheet_name, output_filename,
                 initial_row, next_row_range, initial_column, next_column_block_range, column_block_number):
        patch_worksheet()
        try:
            self.wb = load_workbook(settings.REPORT_TEMPLATE_DIRECTORY + template_name)
            self.sheet = self.wb[sheet_name]
        except Exception as e:
            print(e)
            raise APIException("Error in excel report generation -> %s" % e)

        # file name böyle olmaz. isim yazdırmak için düzgün bi algoritma yazmak gerek
        self.output_filename = output_filename

        self.current_row = initial_row
        self.row_range = next_row_range

        self.initial_column = initial_column
        self.current_column = initial_column
        self.column_block_range = next_column_block_range
        self.column_block_number = column_block_number
        self.column_count = 1

    def write_current_row(self, column, item):
        """current row ve paramtre olarak verilen column'a yazar. başka işlem yapmaz"""
        self.sheet.cell(row=self.current_row, column=column).value = item

    def write_current_position(self, item):
        """current row ve column'a yazar ve column'u 1 arttırır"""
        self.sheet.cell(row=self.current_row, column=self.current_column).value = item
        self.current_column += 1

    def next_column_block(self):
        """
        kolonlar gruplanmış durumda. kolonlar arka arkaya ve gruplar arası boşluk var.
        verilen column_block_number kadar ilerledikten sonra başa döner. daktilo gibi ama aşağı satıra inmez ;)
        :return:
        """

        if self.column_count == self.column_block_number:
            self.column_count = 1
            self.current_column = self.initial_column
        else:
            self.column_count += 1
            self.current_column += self.column_block_range

    def next_row(self):
        """bir alt satıra geçer"""
        self.current_row += self.row_range

    def save(self):
        try:
            self.wb.save(settings.REPORT_DAILYSTUDY_DIRECTORY + self.output_filename)
        except Exception as e:
            print(e)
            raise APIException("Error in excel report save -> %s" % e)
